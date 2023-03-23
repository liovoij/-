from openpyxl import *
from openpyxl.styles import *
from requestconfig.use_config import *


class ParsExcel(object):
    # 实例化，定位excel
    def __init__(self):
        self.wb = load_workbook(xlsx_dir)
        self.sheetnames = self.wb.sheetnames[0]
        self.ws = self.wb[self.sheetnames]

    # 获取sheet对象
    def SheetName(self, sheetname):
        sheet = self.wb[sheetname]
        return sheet

    # 获取表格的行数
    def get_row(self, sheet):
        rows = sheet.max_row
        return rows

    # 获取表格的列数
    def get_column(self, sheet):
        column = sheet.max_column
        return column

    # 获取表格某一行的数据
    def get_row_value(self, sheet, row_num):
        maxcolumn = self.get_row(sheet)
        row_data = []
        for i in range(1, maxcolumn + 1):
            cell_value = sheet.cell(row=row_num, column=i).value
            if cell_value is None:
                cell_value = ''
            row_data.append(cell_value)
        return row_data

    # 获取表格某一列的数据
    def get_column_value(self, sheet, column_num):
        maxrow = self.get_column(sheet)
        column_data = []
        for i in range(2, maxrow + 1):
            cell_value = sheet.cell(row=i, column=column_num).value
            if cell_value is None:
                cell_value = ''
            column_data.append(cell_value)
        return column_data

    # 获取表格某一单元格的数据
    def get_cell_value(self, sheet, row_num, column_num):
        cell_value = sheet.cell(row_num, column_num).value
        if cell_value is None:
            cell_value = ''
        return cell_value

    # 获取某sheet页的所有测试数据(除去表头)，返回一个元组组成的列表
    def get_values_sheet(self, sheet):
        row_num = self.get_row(sheet)
        column_num = self.get_column(sheet)
        all_values = []
        for row in range(2, row_num + 1):
            value_list = []
            for column in range(1, column_num + 1):
                values = sheet.cell(row, column).value
                if values is None:
                    values = ''
                value_list.append(values)
            all_values.append(all_values)
        return all_values

    def set_cell_value(self, sheet, row, column, text):
        sheet.cell(row, column, value=str(text))

    def right_cell_style(self, sheet, row, column):
        right_12_font = Font(name='宋体', size=12, italic=False, color='000000', bold=True)  # 黑色
        sheet.cell(row, column).font = right_12_font

    def wrong_cell_style(self, sheet, row, column):
        wrong_12_font = Font(name='宋体', size=12, italic=False, color='FF0000', bold=True)  # 红色
        sheet.cell(row, column).font = wrong_12_font

    # 保存execl数据
    def saveData(self):
        self.wb.save(xlsx_dir)


if __name__ == '__main__':
    sheet = ParsExcel().SheetName('通讯工具')
    print(ParsExcel().SheetName('通讯工具'))
    print(ParsExcel().get_row())
    print(ParsExcel().get_column())
